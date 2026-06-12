"""Server-side report generation: turns a JSON "spec" into a real .xlsx workbook
or a paginated PDF.

The frontend already holds every row it needs (and all the filtering / field-
selection UI), so it assembles a spec and POSTs it here purely for rendering:

    {
      "title": "Tasks",
      "sheets": [
        {
          "name": "Tasks",
          "columns": [{"key": "client_name", "label": "Client"}, ...],
          "rows":    [{"client_name": "ACME", ...}, ...]
        },
        ...
      ]
    }

Each sheet becomes a worksheet (Excel) or a titled table section (PDF). Nothing
sensitive is ever rendered here that the caller didn't already send — vault
passwords are never part of any spec.
"""
from io import BytesIO

# Defensive caps so a malformed/oversized spec can't exhaust memory.
MAX_SHEETS = 25
MAX_ROWS = 50000
MAX_COLS = 60


class ExportError(Exception):
    """Bad export spec -> HTTP 400."""
    def __init__(self, message):
        self.message = message
        super().__init__(message)


def _coerce(value):
    """Normalise a cell value for display."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def normalise_spec(spec):
    """Validate and trim a raw spec dict to safe, well-formed sheets."""
    if not isinstance(spec, dict):
        raise ExportError("Export body must be a JSON object.")
    title = str(spec.get("title") or "Report").strip() or "Report"
    sheets_in = spec.get("sheets")
    if not isinstance(sheets_in, list) or not sheets_in:
        raise ExportError("Export spec must include at least one sheet.")
    if len(sheets_in) > MAX_SHEETS:
        raise ExportError(f"Too many sheets (max {MAX_SHEETS}).")

    sheets = []
    for s in sheets_in:
        if not isinstance(s, dict):
            continue
        name = str(s.get("name") or "Sheet").strip() or "Sheet"
        columns = s.get("columns") or []
        rows = s.get("rows") or []
        if not isinstance(columns, list) or not isinstance(rows, list):
            raise ExportError(f"Sheet '{name}' has an invalid columns/rows structure.")
        if len(columns) > MAX_COLS:
            raise ExportError(f"Sheet '{name}' has too many columns (max {MAX_COLS}).")
        if len(rows) > MAX_ROWS:
            raise ExportError(f"Sheet '{name}' has too many rows (max {MAX_ROWS}).")
        cols = []
        for c in columns:
            if isinstance(c, dict) and c.get("key") is not None:
                col = {"key": str(c["key"]), "label": str(c.get("label") or c["key"])}
                # Optional dropdown list for this column (import templates).
                if isinstance(c.get("options"), list) and c["options"]:
                    col["options"] = [str(o) for o in c["options"]][:40]
                cols.append(col)
        if not cols:
            # Fall back to keys discovered on the first row
            if rows and isinstance(rows[0], dict):
                cols = [{"key": str(k), "label": str(k)} for k in rows[0].keys()]
            else:
                raise ExportError(f"Sheet '{name}' has no columns to export.")
        sheets.append({"name": name, "columns": cols, "rows": rows})
    if not sheets:
        raise ExportError("Export spec produced no usable sheets.")
    return {"title": title, "sheets": sheets, "no_title": bool(spec.get("no_title"))}


# --------------------------------------------------------------------------- #
#  Excel (.xlsx) via openpyxl
# --------------------------------------------------------------------------- #
def build_xlsx(spec):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    spec = normalise_spec(spec)
    # `no_title` (used by import templates) puts the column headers on row 1
    # instead of a merged section-title banner on row 1 + headers on row 2.
    no_title = bool(spec.get("no_title"))
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    header_fill = PatternFill("solid", fgColor="1C1C1C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1C1C1C")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(vertical="top", wrap_text=True)

    # Hidden helper sheet that holds dropdown option lists. Referencing a range
    # here (instead of an inline "a,b,c" formula) sidesteps Excel's ~255-char
    # limit on list validations, so dropdowns work for long client/service lists.
    from openpyxl.utils import get_column_letter as _gcl
    lists_ws = None
    lists_col = 0

    def _list_ref(options):
        nonlocal lists_ws, lists_col
        if lists_ws is None:
            lists_ws = wb.create_sheet(title="_lists")
            lists_ws.sheet_state = "hidden"
        lists_col += 1
        letter = _gcl(lists_col)
        for ri, opt in enumerate(options, start=1):
            lists_ws.cell(row=ri, column=lists_col, value=str(opt))
        return f"=_lists!${letter}$1:${letter}${len(options)}"

    used_names = set()
    for sheet in spec["sheets"]:
        # Excel sheet titles: <=31 chars, no  : \ / ? * [ ]  and must be unique
        raw = sheet["name"][:31]
        for ch in r':\/?*[]':
            raw = raw.replace(ch, " ")
        base = raw.strip() or "Sheet"
        title = base
        n = 2
        while title.lower() in used_names:
            suffix = f" ({n})"
            title = base[:31 - len(suffix)] + suffix
            n += 1
        used_names.add(title.lower())

        ws = wb.create_sheet(title=title)
        columns = sheet["columns"]
        ncols = len(columns)

        if no_title:
            header_row = 1
        else:
            # Row 1: section title spanning all columns
            ws.cell(row=1, column=1, value=sheet["name"]).font = title_font
            if ncols > 1:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            header_row = 2

        # Column headers
        for ci, col in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=ci, value=col["label"])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

        # Track max content width per column for auto-sizing
        widths = [len(str(col["label"])) for col in columns]

        for ri, row in enumerate(sheet["rows"], start=header_row + 1):
            for ci, col in enumerate(columns, start=1):
                val = _coerce(row.get(col["key"]) if isinstance(row, dict) else "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = wrap_top
                cell.border = border
                text_len = len(str(val))
                if text_len > widths[ci - 1]:
                    widths[ci - 1] = text_len

        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 10), 55)

        # Per-column dropdown lists (e.g. Entity Type on the clients template,
        # Service / Client on the tasks template). Applied to a generous range
        # of data rows below the header. Short lists use an inline formula; long
        # ones are written to the hidden _lists sheet and referenced by range.
        first_data = header_row + 1
        last_data = first_data + 500
        for ci, col in enumerate(columns, start=1):
            opts = col.get("options") if isinstance(col, dict) else None
            if not opts:
                continue
            letter = get_column_letter(ci)
            inline = '"' + ",".join(str(o).replace(",", " ") for o in opts) + '"'
            formula = inline if len(inline) <= 250 else _list_ref(opts)
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.add(f"{letter}{first_data}:{letter}{last_data}")
            ws.add_data_validation(dv)

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        if sheet["rows"]:
            ws.auto_filter.ref = (
                f"{get_column_letter(1)}{header_row}:"
                f"{get_column_letter(ncols)}{header_row + len(sheet['rows'])}"
            )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# --------------------------------------------------------------------------- #
#  PDF via reportlab (platypus)
# --------------------------------------------------------------------------- #
def build_pdf(spec):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )

    spec = normalise_spec(spec)
    buf = BytesIO()
    page_size = landscape(A4)
    left = right = 12 * mm
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=left, rightMargin=right, topMargin=12 * mm, bottomMargin=12 * mm,
        title=spec["title"],
    )
    avail_width = page_size[0] - left - right

    styles = getSampleStyleSheet()
    h_style = ParagraphStyle("secTitle", parent=styles["Heading2"],
                             textColor=colors.HexColor("#1C1C1C"), spaceAfter=6)
    doc_title = ParagraphStyle("docTitle", parent=styles["Title"],
                               fontSize=18, textColor=colors.HexColor("#1C1C1C"), spaceAfter=2)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=8, textColor=colors.HexColor("#888888"), spaceAfter=10)

    story = [Paragraph(spec["title"], doc_title)]
    from datetime import datetime
    story.append(Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))

    for si, sheet in enumerate(spec["sheets"]):
        columns = sheet["columns"]
        ncols = max(1, len(columns))
        # Shrink font as the table gets wider so it still fits the page.
        font_size = 8 if ncols <= 8 else (7 if ncols <= 12 else 6)
        cell_style = ParagraphStyle("cell", parent=styles["Normal"],
                                    fontSize=font_size, leading=font_size + 2)
        head_style = ParagraphStyle("head", parent=styles["Normal"],
                                    fontSize=font_size, leading=font_size + 2,
                                    textColor=colors.white, fontName="Helvetica-Bold")

        if si > 0:
            story.append(Spacer(1, 6 * mm))
        story.append(Paragraph(sheet["name"], h_style))

        header = [Paragraph(str(c["label"]), head_style) for c in columns]
        data = [header]
        for row in sheet["rows"]:
            line = []
            for c in columns:
                val = _coerce(row.get(c["key"]) if isinstance(row, dict) else "")
                line.append(Paragraph(str(val).replace("\n", "<br/>"), cell_style))
            data.append(line)

        if not sheet["rows"]:
            data.append([Paragraph("<i>No records.</i>", cell_style)] + [""] * (ncols - 1))

        col_width = avail_width / ncols
        table = Table(data, colWidths=[col_width] * ncols, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1C")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F6F6")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def render(spec, fmt):
    """Dispatch to the requested format. Returns (bytes, mimetype, extension)."""
    fmt = (fmt or "xlsx").lower()
    if fmt == "pdf":
        return build_pdf(spec), "application/pdf", "pdf"
    if fmt in ("xlsx", "excel"):
        return (build_xlsx(spec),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "xlsx")
    raise ExportError(f"Unsupported export format '{fmt}'. Use 'xlsx' or 'pdf'.")
