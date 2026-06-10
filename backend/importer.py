"""Bulk import: parse an uploaded CSV or Excel file into normalised rows, with
flexible header matching. Supports several entity types (clients, services,
users, tasks).

Per-entity validation and database insertion live in main.py / crud.py; this
module only turns a spreadsheet into clean dict rows. Unknown columns are ignored.
"""
import csv
import io

MAX_IMPORT_ROWS = 10000


class ImportFileError(Exception):
    """Unreadable / unsupported / wrong-shaped upload -> HTTP 400."""
    def __init__(self, message):
        self.message = message
        super().__init__(message)


# Each entity defines: header aliases (canonical -> accepted spellings),
# the template columns shown to the user (key, friendly label, required?),
# and a sample row for the downloadable template.
ENTITY_CONFIGS = {
    "clients": {
        "label": "Clients",
        "aliases": {
            "name": ["name", "client name", "client full name", "clientname", "fullname"],
            "entity_type": ["entity type", "entity", "type", "constitution"],
            "pan": ["pan", "pan no", "pan number", "pan card"],
            "gstin": ["gstin", "gst", "gst no", "gst number", "gstin number"],
            "group": ["group", "client group", "family", "group name", "client group family"],
            "physical_folder_location": ["folder location", "physical folder location", "cabinet",
                                         "folder", "physical folder"],
            "data_location": ["data location", "digital location", "data path", "softcopy location"],
        },
        "template_columns": [
            ("name", "Name", True),
            ("entity_type", "Entity Type", False),
            ("pan", "PAN", False),
            ("gstin", "GSTIN", False),
            ("group", "Group", False),
            ("physical_folder_location", "Folder Location", False),
            ("data_location", "Data Location", False),
        ],
        "sample": {
            "name": "ACME Industries Pvt Ltd", "entity_type": "Company", "pan": "AAACA1234A",
            "gstin": "27AAACA1234A1Z5", "group": "ACME Group",
            "physical_folder_location": "Cabinet 3 / Shelf B", "data_location": "D:/clients/acme",
        },
    },
    "services": {
        "label": "Services",
        "aliases": {
            "name": ["name", "service", "service name", "service template"],
            "description": ["description", "desc", "notes", "about"],
            "checklist": ["checklist", "checklist steps", "steps", "checklist items"],
            "default_due_day": ["default due day", "due day", "due date day"],
        },
        "template_columns": [
            ("name", "Name", True),
            ("description", "Description", False),
            ("checklist", "Checklist", False),
            ("default_due_day", "Default Due Day", False),
        ],
        "sample": {
            "name": "GST Return Filing", "description": "Monthly GSTR-3B preparation and filing",
            "checklist": "Collect invoices, Reconcile 2B, File return", "default_due_day": "20",
        },
    },
    "users": {
        "label": "Staff Users",
        "aliases": {
            "full_name": ["full name", "name", "staff name", "employee name", "fullname"],
            "username": ["username", "user name", "login", "user id", "userid"],
            "password": ["password", "pass", "pwd"],
        },
        "template_columns": [
            ("full_name", "Full Name", True),
            ("username", "Username", True),
            ("password", "Password", True),
        ],
        "sample": {"full_name": "Riya Mehta", "username": "riya", "password": "ChangeMe@123"},
    },
    "tasks": {
        "label": "Tasks",
        "aliases": {
            "client": ["client", "client name", "client full name"],
            "service": ["service", "service name", "service template"],
            "financial_year": ["financial year", "fy", "year"],
            "period": ["period", "periodicity", "month", "quarter"],
            "status": ["status", "state"],
            "assigned_to": ["assigned to", "assigned staff", "staff", "assignee"],
            "due_date": ["due date", "due", "deadline"],
        },
        "template_columns": [
            ("client", "Client", True),
            ("service", "Service", True),
            ("financial_year", "Financial Year", True),
            ("period", "Period", True),
            ("status", "Status", False),
            ("assigned_to", "Assigned To", False),
            ("due_date", "Due Date", False),
        ],
        "sample": {
            "client": "ACME Industries Pvt Ltd", "service": "Income Tax Return",
            "financial_year": "2025-26", "period": "Annual", "status": "Working",
            "assigned_to": "Riya Mehta", "due_date": "2026-03-31",
        },
    },
}


def _norm_header(h):
    return " ".join(str(h or "").strip().lower().replace("_", " ").replace("/", " ").split())


def _build_header_map(headers, aliases):
    """Map source column index -> canonical field name (or omit)."""
    lookup = {}
    for canon, names in aliases.items():
        for a in names:
            lookup[_norm_header(a)] = canon
        lookup[_norm_header(canon)] = canon
    mapping = {}
    for idx, h in enumerate(headers):
        canon = lookup.get(_norm_header(h))
        if canon and canon not in mapping.values():
            mapping[idx] = canon
    return mapping


def _rows_from_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    return [list(r) for r in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(data: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if c is None else c for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def parse_upload(filename: str, data: bytes, entity: str):
    """Return a list of dicts {canonical_field: value, "_row": <1-based row no>}.

    Raises ImportFileError for unsupported/empty/headerless files or when a
    required column is missing.
    """
    if entity not in ENTITY_CONFIGS:
        raise ImportFileError(f"Unknown import type '{entity}'.")
    cfg = ENTITY_CONFIGS[entity]

    name = (filename or "").lower()
    if name.endswith(".csv"):
        grid = _rows_from_csv(data)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        grid = _rows_from_xlsx(data)
    else:
        raise ImportFileError("Unsupported file type. Upload a .csv or .xlsx file.")

    grid = [row for row in grid if any(str(c).strip() for c in row)]
    if not grid:
        raise ImportFileError("The file is empty.")

    headers = grid[0]
    mapping = _build_header_map(headers, cfg["aliases"])

    label_of = {k: lbl for k, lbl, _req in cfg["template_columns"]}
    required_keys = [k for k, _lbl, req in cfg["template_columns"] if req]
    missing = [label_of.get(k, k) for k in required_keys if k not in mapping.values()]
    if missing:
        raise ImportFileError(
            "The header row is missing required column(s): " + ", ".join(missing) +
            ". The first row of the file must name the columns."
        )

    body = grid[1:]
    if len(body) > MAX_IMPORT_ROWS:
        raise ImportFileError(f"Too many rows (max {MAX_IMPORT_ROWS}).")

    out = []
    for i, row in enumerate(body, start=2):
        rec = {"_row": i}
        for idx, canon in mapping.items():
            val = row[idx] if idx < len(row) else ""
            rec[canon] = str(val).strip() if val is not None else ""
        if not any(v for k, v in rec.items() if k != "_row"):
            continue
        out.append(rec)
    return out
